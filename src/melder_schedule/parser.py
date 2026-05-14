from __future__ import annotations

import hashlib
import math
import re
from pathlib import Path
from typing import Iterable

import openpyxl

from melder_schedule.models import ActivityType, LessonRequest


LECTURE_COLUMN = 6
PRACTICE_COLUMN = 7
GROUP_COLUMN = 5
SUBJECT_COLUMN = 2
KIND_COLUMN = 3

DEFAULT_WEEKS_IN_SEMESTER = 17
ACADEMIC_HOURS_IN_PAIR = 2


def parse_workload_excel(
    path: str | Path,
    *,
    sheet_name: str | None = None,
    include_activity_types: Iterable[ActivityType] = (
        ActivityType.LECTURE,
        ActivityType.PRACTICE,
    ),
) -> list[LessonRequest]:
    """Parse university workload Excel into normalized lesson requests.

    The workload file stores semester hours. For a weekly timetable prototype we
    convert them into a number of pairs in a typical week: 34 hours in a
    17-week semester equals one weekly pair.
    """

    include = set(include_activity_types)
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook[sheet_name] if sheet_name else workbook[workbook.sheetnames[0]]

    requests: list[LessonRequest] = []
    current_teacher = ""
    current_semester = ""
    current_subject = ""

    for row_idx, row in enumerate(worksheet.iter_rows(values_only=True), start=1):
        values = [clean_cell(value) for value in row]
        if not any(values):
            continue

        number = values[0] if len(values) > 0 else ""
        title = values[SUBJECT_COLUMN - 1] if len(values) >= SUBJECT_COLUMN else ""
        kind = values[KIND_COLUMN - 1] if len(values) >= KIND_COLUMN else ""
        groups_cell = values[GROUP_COLUMN - 1] if len(values) >= GROUP_COLUMN else ""

        if is_teacher_row(number, title):
            current_teacher = title
            current_semester = ""
            current_subject = ""
            continue

        if is_semester_row(title):
            current_semester = title
            current_subject = ""
            continue

        if kind in {"Группа", "Поток"} and groups_cell and current_teacher:
            if title:
                current_subject = title
            if not current_subject:
                continue

            groups = split_groups(groups_cell)
            if not groups:
                continue

            lecture_hours = to_float(values[LECTURE_COLUMN - 1] if len(values) >= LECTURE_COLUMN else "")
            practice_hours = to_float(values[PRACTICE_COLUMN - 1] if len(values) >= PRACTICE_COLUMN else "")

            if ActivityType.LECTURE in include and lecture_hours > 0:
                requests.append(
                    build_request(
                        row_idx=row_idx,
                        teacher=current_teacher,
                        semester=current_semester,
                        subject=current_subject,
                        activity_type=ActivityType.LECTURE,
                        groups=groups,
                        source_kind=kind,
                        hours=lecture_hours,
                    )
                )

            if ActivityType.PRACTICE in include and practice_hours > 0:
                requests.append(
                    build_request(
                        row_idx=row_idx,
                        teacher=current_teacher,
                        semester=current_semester,
                        subject=current_subject,
                        activity_type=ActivityType.PRACTICE,
                        groups=groups,
                        source_kind=kind,
                        hours=practice_hours,
                    )
                )

    return requests


def build_request(
    *,
    row_idx: int,
    teacher: str,
    semester: str,
    subject: str,
    activity_type: ActivityType,
    groups: tuple[str, ...],
    source_kind: str,
    hours: float,
) -> LessonRequest:
    required_slots = hours_to_weekly_slots(hours)
    room_type = "lecture" if activity_type == ActivityType.LECTURE or source_kind == "Поток" else "classroom"
    estimated_students = estimate_students(groups, source_kind)
    identity = "|".join(
        [
            str(row_idx),
            teacher,
            semester,
            subject,
            activity_type.value,
            ",".join(groups),
            str(hours),
        ]
    )
    request_id = hashlib.sha1(identity.encode("utf-8")).hexdigest()[:12]
    return LessonRequest(
        id=request_id,
        teacher=teacher,
        semester=semester or "Не указан",
        subject=subject,
        activity_type=activity_type,
        groups=groups,
        source_kind=source_kind,
        hours=hours,
        required_slots=required_slots,
        required_room_type=room_type,
        estimated_students=estimated_students,
    )


def hours_to_weekly_slots(hours: float) -> int:
    hours_per_weekly_pair = DEFAULT_WEEKS_IN_SEMESTER * ACADEMIC_HOURS_IN_PAIR
    return max(1, math.ceil(hours / hours_per_weekly_pair))


def clean_cell(value: object) -> str:
    if value is None:
        return ""
    text = str(value).replace("\xa0", " ").strip()
    return re.sub(r"\s+", " ", text)


def to_float(value: object) -> float:
    if value in (None, ""):
        return 0.0
    try:
        return float(str(value).replace(",", "."))
    except ValueError:
        return 0.0


def is_teacher_row(number: str, title: str) -> bool:
    return bool(re.fullmatch(r"\d+\.?", number.strip())) and bool(title)


def is_semester_row(title: str) -> bool:
    return "семестр" in title.lower()


def split_groups(value: str) -> tuple[str, ...]:
    parts = [part.strip() for part in re.split(r"[,;]", value) if part.strip()]
    return tuple(dict.fromkeys(parts))


def estimate_students(groups: tuple[str, ...], source_kind: str) -> int:
    if source_kind == "Поток":
        return min(180, max(30, len(groups) * 25))
    return min(180, max(15, len(groups) * 25))
